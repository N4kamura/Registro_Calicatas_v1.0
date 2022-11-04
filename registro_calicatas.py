from pickle import TRUE
import xlsxwriter
import openpyxl

#Lectura de datos de otro excel datos_calicatas.xlsx
path="C:\\Users\\HP\\Desktop\\DARYL_SUELOS\\Programas\\datos_calicatas.xlsx"

wb_obj=openpyxl.load_workbook(path)

sheet_obj=wb_obj.active

cell_obj=sheet_obj.cell(row=2,column=4)
proyecto=cell_obj.value

cell_obj=sheet_obj.cell(row=3,column=4)
ubicacion=cell_obj.value

cell_obj=sheet_obj.cell(row=4,column=4)
tipo_exploracion=cell_obj.value

workbook = xlsxwriter.Workbook('C:\\Users\\HP\\Desktop\\DARYL_SUELOS\\Programas\\Registros_Calicatas_V2.xlsx')

#EL BUCLE MÁXIMO
k=0
cantidad_calicatas=21
for k in range(cantidad_calicatas):

    fila=7+k
    columna=1
    #Calicata
    cell_obj=sheet_obj.cell(row=fila,column=columna)
    calicata=cell_obj.value

    #Nivel freatico
    cell_obj=sheet_obj.cell(row=fila,column=columna+1)
    nivel_freatico=cell_obj.value

    #Prof. de exploracion
    cell_obj=sheet_obj.cell(row=fila,column=columna+2)
    profundidad_exploracion=cell_obj.value

    #Este
    cell_obj=sheet_obj.cell(row=fila,column=columna+3)
    este=cell_obj.value

    #Norte
    cell_obj=sheet_obj.cell(row=fila,column=columna+4)
    norte=cell_obj.value

    #Fecha
    cell_obj=sheet_obj.cell(row=fila,column=columna+5)
    fecha=cell_obj.value

    #Profundidad 1
    cell_obj=sheet_obj.cell(row=fila,column=columna+6)
    profundidad1=cell_obj.value

    #Profundidad 2
    cell_obj=sheet_obj.cell(row=fila,column=columna+7)
    profundidad2=cell_obj.value

    #Profundidad 3
    cell_obj=sheet_obj.cell(row=fila,column=columna+8)
    profundidad3=cell_obj.value

    #SUCS 1
    cell_obj=sheet_obj.cell(row=fila,column=columna+9)
    sucs1=cell_obj.value

    #SUCS 2
    cell_obj=sheet_obj.cell(row=fila,column=columna+10)
    sucs2=cell_obj.value

    #SUCS 3
    cell_obj=sheet_obj.cell(row=fila,column=columna+11)
    sucs3=cell_obj.value

    #MUESTRA 1
    cell_obj=sheet_obj.cell(row=fila,column=columna+12)
    muestra1=cell_obj.value

    #MUESTRA 2
    cell_obj=sheet_obj.cell(row=fila,column=columna+13)
    muestra2=cell_obj.value

    #MUESTRA 3
    cell_obj=sheet_obj.cell(row=fila,column=columna+14)
    muestra3=cell_obj.value

    #DESCRIPCION ESTRATIGRAFICA 1
    cell_obj=sheet_obj.cell(row=fila,column=columna+15)
    descripcion_estratigrafica1=cell_obj.value

    #DESCRIPCION ESTRATIGRAFICA 2
    cell_obj=sheet_obj.cell(row=fila,column=columna+16)
    descripcion_estratigrafica2=cell_obj.value

    #DESCRIPCION ESTRATIGRAFICA 3
    cell_obj=sheet_obj.cell(row=fila,column=columna+17)
    descripcion_estratigrafica3=cell_obj.value

    #GRANULOMETRIA 1
    cell_obj=sheet_obj.cell(row=fila,column=columna+18)
    granulometria1=cell_obj.value

    #GRANULOMETRIA 2
    cell_obj=sheet_obj.cell(row=fila,column=columna+19)
    granulometria2=cell_obj.value

    #GRANULOMETRIA 3
    cell_obj=sheet_obj.cell(row=fila,column=columna+20)
    granulometria3=cell_obj.value

    #INDICE PLASTICO
    cell_obj=sheet_obj.cell(row=fila,column=columna+21)
    indice_plastico=cell_obj.value

    #%HUMEDAD
    cell_obj=sheet_obj.cell(row=fila,column=columna+22)
    humedad=cell_obj.value

    ##############################################

    # Creamos hoja y libro excel
    worksheet = workbook.add_worksheet(calicata)

    #Bordes
    cell_format=workbook.add_format({'border':True})

    #Formatos
    style_negrita=workbook.add_format({'bold':True,'left':1,'font_size':8})
    style_negrita2=workbook.add_format({'bold':True,'font_size':8,'align':'center'})
    style_encabezado=workbook.add_format({'border':1,'align':'center', 'bold':True})
    style_encabezado2=workbook.add_format({'border':1,'text_wrap':True,'bold':True,'font_size':9,'valign':'vcenter','align':'center'})
    style_encabezado3=workbook.add_format({'border':1,'text_wrap':True,'bold':True,'font_size':7,'valign':'vcenter','align':'center'})
    style_encabezado4=workbook.add_format({'border':1,'text_wrap':True,'bold':True,'font_size':9,'valign':'vcenter','align':'center','rotation':90})
    style_fotos=workbook.add_format({'border':1})
    style_PROYECTO=workbook.add_format({'left':1,'top':1,'bold':True,'font_size':8})
    style_FECHA=workbook.add_format({'left':1,'bottom':1,'bold':True,'font_size':8})
    style_PRIMERDOSPTOS=workbook.add_format({'top':1,'font_size':8,'bold':True,'align':'center'})
    style_ULTIMODOSPTOS=workbook.add_format({'bottom':1,'font_size':8,'bold':True,'align':'center'})
    style_INTRODUCIRNOMBRE=workbook.add_format({'top':1,'right':1,'font_size':8})
    style_INTRODUCIRFECHA=workbook.add_format({'bottom':1,'right':1,'font_size':8})
    style_INTRODUCIR=workbook.add_format({'right':1,'font_size':8})
    style_descripcion=workbook.add_format({'text_wrap':True,'border':1,'valign':'vcenter','align':'center','font_size':8})
    style_simbolo=workbook.add_format({'border':1,'valign':'vcenter','align':'center','bold':True})
    style_simbolo2=workbook.add_format({'border':1,'valign':'bottom','align':'center','bold':True})
    style_SUCS=workbook.add_format({'border':1,'valign':'vcenter','align':'center','bold':True,'font_size':8})
    style_profundo=workbook.add_format({'bottom':1})
    style_napafreatica=workbook.add_format({'bg_color':'cyan'})
    style_profundo2=workbook.add_format({'bottom':1,'bg_color':'cyan'})
    style_lateral=workbook.add_format({'border':1,'bg_color':'black'})

    #Ajustando columnas
    worksheet.set_column(0,0,0.67) #A
    worksheet.set_column(1,1,1.29) #B
    worksheet.set_column(2,2,5.57) #C
    worksheet.set_column(3,3,9.43) #D
    worksheet.set_column(4,4,1.14) #E
    worksheet.set_column(5,5,27.86) #F
    worksheet.set_column(6,6,4.71) #G
    worksheet.set_column(7,7,5) #H
    worksheet.set_column(8,8,5) #I
    worksheet.set_column(9,9,5) #J
    worksheet.set_column(10,10,5) #K
    worksheet.set_column(11,11,5) #L
    worksheet.set_column(12,12,5.43) #M

    #Ajustando filas
    worksheet.set_row(0,13.5) #1
    worksheet.set_row(1,5) #2
    i=0
    for i in range(8):
        worksheet.set_row(i+2,12) #3-10
    worksheet.set_row(10,5) #11
    i=0
    for i in range(10):
        worksheet.set_row(i+11,12) #12 - 21
    worksheet.set_row(21,5) #22
    worksheet.set_row(22,17.25) #23
    worksheet.set_row(23,27.75) #24
    i=0
    for i in range(40):
        worksheet.set_row(i+24,12)
    worksheet.set_row(64,5) #65
    i=0
    for i in range(3):
        worksheet.set_row(i+65,11.5)

    # Combinando filas
    worksheet.merge_range(0,0,0,12,'REGISTRO DE EXCAVACION DE CALICATA',style_encabezado)
    worksheet.merge_range(2,0,2,3,'PROYECTO',style_PROYECTO)
    worksheet.merge_range(2,5,2,12,proyecto,style_INTRODUCIRNOMBRE)
    worksheet.merge_range(3,0,3,3,'UBICACIÓN',style_negrita)
    worksheet.merge_range(3,5,3,12,ubicacion,style_INTRODUCIR)
    worksheet.merge_range(4,0,4,3,'TIPO DE EXPLORACION',style_negrita)
    worksheet.merge_range(4,5,4,12,tipo_exploracion,style_INTRODUCIR)
    worksheet.merge_range(5,0,5,3,'N° DE EXPLORACION',style_negrita)
    worksheet.merge_range(5,5,5,12,calicata,style_INTRODUCIR)
    worksheet.merge_range(6,0,6,3,'NIVEL FREATICO (m)',style_negrita)
    if nivel_freatico==0:
        worksheet.merge_range(6,5,6,12,"No Presenta",style_INTRODUCIR)
    else:
        worksheet.merge_range(6,5,6,12,"{:.2f}".format(nivel_freatico),style_INTRODUCIR)
    worksheet.merge_range(7,0,7,3,'PROF. DE EXPLORACION (m)',style_negrita)
    worksheet.merge_range(7,5,7,12,"{:.2f}".format(profundidad_exploracion),style_INTRODUCIR)
    worksheet.merge_range(8,0,8,3,'COORDENADAS UTM',style_negrita)
    worksheet.merge_range(8,5,8,12,str(este)+'mE '+str(norte)+'mN',style_INTRODUCIR)
    worksheet.merge_range(9,0,9,3,'FECHA DE EXCAVACION',style_FECHA)
    worksheet.merge_range(9,5,9,12,str(fecha),style_INTRODUCIRFECHA)
    worksheet.write('E3',':',style_PRIMERDOSPTOS)
    worksheet.write('E4',':',style_negrita2)
    worksheet.write('E5',':',style_negrita2)
    worksheet.write('E6',':',style_negrita2)
    worksheet.write('E7',':',style_negrita2)
    worksheet.write('E8',':',style_negrita2)
    worksheet.write('E9',':',style_negrita2)
    worksheet.write('E10',':',style_ULTIMODOSPTOS)

    #Dibujo.
    worksheet.merge_range(11,0,20,12,'',style_fotos)

    #Encabezado
    worksheet.merge_range(22,0,23,2,'PROF. (m)',style_encabezado2)
    worksheet.merge_range(22,3,23,3,'SIMBOLOGIA',style_encabezado2)
    worksheet.merge_range(22,4,23,5,'DESCRIPCION ESTRATIGRAFICA',style_encabezado2)
    worksheet.merge_range(22,6,23,6,'SUCS',style_encabezado4)
    worksheet.merge_range(22,7,22,11,'GRANULOMETRIA',style_encabezado2)
    worksheet.merge_range(22,12,23,12,'N° DE MUESTRA',style_encabezado4)
    worksheet.write('H24','<0.0075mm',style_encabezado3)
    worksheet.write('I24','0.0075mm a 4.750mm',style_encabezado3)
    worksheet.write('J24','4.750mm a 75mm',style_encabezado3)
    worksheet.write('K24','I.P',style_encabezado3)
    worksheet.write('L24','%Hum.',style_encabezado3)

    #Lateral
    i=0
    for i in range(40):
        if i%2==0:
            worksheet.write(24+i,0,'',style_fotos)
        else:
            worksheet.write(24+i,0,'',style_lateral)

    #Estrato
    x1=profundidad1 #Profundidad 1
    x2=profundidad2 #Profundidad 2
    x3=profundidad3 #Profundidad 3

    conta1=int(x1*10) #8
    conta2=int(x2*10) #14
    conta3=int(x3*10) #21

    worksheet.merge_range(24,2,24+conta1-1,2,"{:.2f}".format(profundidad1),style_simbolo2)
    worksheet.merge_range(24,3,24+conta1-1,3,'',style_simbolo)
    worksheet.merge_range(24,4,24+conta1-1,5,descripcion_estratigrafica1,style_descripcion)
    worksheet.merge_range(24,6,24+conta1-1,6,sucs1,style_SUCS)
    worksheet.merge_range(24,12,24+conta1-1,12,muestra1,style_SUCS)

    worksheet.merge_range(24+conta1,2,24+conta2-1,2,"{:.2f}".format(profundidad2),style_simbolo2)
    worksheet.merge_range(24+conta1,3,24+conta2-1,3,'',style_simbolo)
    worksheet.merge_range(24+conta1,4,24+conta2-1,5,descripcion_estratigrafica2,style_descripcion)
    worksheet.merge_range(24+conta1,6,24+conta2-1,6,sucs2,style_SUCS)
    worksheet.merge_range(24+conta1,12,24+conta2-1,12,muestra2,style_SUCS)

    worksheet.merge_range(24+conta2,2,24+conta3-1,2,"{:.2f}".format(profundidad3),style_simbolo2)
    worksheet.merge_range(24+conta2,3,24+conta3-1,3,'',style_simbolo)
    worksheet.merge_range(24+conta2,4,24+conta3-1,5,descripcion_estratigrafica3,style_descripcion)
    worksheet.merge_range(24+conta2,6,24+conta3-1,6,sucs3,style_SUCS)
    worksheet.merge_range(24+conta2,12,24+conta3-1,12,muestra3,style_SUCS)

    #Solo para la muestra sacada:
    if muestra1=="M-1":
        worksheet.merge_range(24,7,24+conta1-1,7,"{:.2f}".format(granulometria1),style_SUCS)
        worksheet.merge_range(24,8,24+conta1-1,8,"{:.2f}".format(granulometria2),style_SUCS)
        worksheet.merge_range(24,9,24+conta1-1,9,"{:.2f}".format(granulometria3),style_SUCS)
        worksheet.merge_range(24,10,24+conta1-1,10,indice_plastico,style_SUCS)
        worksheet.merge_range(24,11,24+conta1-1,11,"{:.2f}".format(humedad),style_SUCS)
    else:
        i=0
        for i in range(5):
            worksheet.merge_range(24,7+i,24+conta1-1,7+i,'-',style_SUCS)
    

    if muestra2=="M-1":
        worksheet.merge_range(24+conta1,7,24+conta2-1,7,"{:.2f}".format(granulometria1),style_SUCS)
        worksheet.merge_range(24+conta1,8,24+conta2-1,8,"{:.2f}".format(granulometria2),style_SUCS)
        worksheet.merge_range(24+conta1,9,24+conta2-1,9,"{:.2f}".format(granulometria3),style_SUCS)
        worksheet.merge_range(24+conta1,10,24+conta2-1,10,indice_plastico,style_SUCS)
        worksheet.merge_range(24+conta1,11,24+conta2-1,11,"{:.2f}".format(humedad),style_SUCS)
    else:
        i=0
        for i in range(5):
            worksheet.merge_range(24+conta1,7+i,24+conta2-1,7+i,'-',style_SUCS)

    if muestra3=="M-1":
        worksheet.merge_range(24+conta2,7,24+conta3-1,7,"{:.2f}".format(granulometria1),style_SUCS)
        worksheet.merge_range(24+conta2,8,24+conta3-1,8,"{:.2f}".format(granulometria2),style_SUCS)
        worksheet.merge_range(24+conta2,9,24+conta3-1,9,"{:.2f}".format(granulometria3),style_SUCS)
        worksheet.merge_range(24+conta2,10,24+conta3-1,10,indice_plastico,style_SUCS)
        worksheet.merge_range(24+conta2,11,24+conta3-1,11,"{:.2f}".format(humedad),style_SUCS)
    else:
        i=0
        for i in range(5):
            worksheet.merge_range(24+conta2,7+i,24+conta3-1,7+i,'-',style_SUCS)


    #Nivel freatico
    np=nivel_freatico #profundidad de la napa freatica
    conta_np=int(np*10)
    if np>0:
        i=1
        for i in range(40-conta_np-1):
            worksheet.write(62-i,1,'',style_napafreatica)
        worksheet.write(63,1,'',style_profundo2)
    else:
        worksheet.write('B64','',style_profundo)

    #Observaciones:
    worksheet.write('A66','Observaciones')

    if profundidad3<4:
        worksheet.merge_range(24+conta3,2,63,2,'',style_fotos)
        worksheet.merge_range(24+conta3,3,63,3,'',style_fotos)
        worksheet.merge_range(24+conta3,4,63,5,'',style_fotos)
        worksheet.merge_range(24+conta3,6,63,6,'',style_fotos)
        worksheet.merge_range(24+conta3,7,63,7,'',style_fotos)
        worksheet.merge_range(24+conta3,8,63,8,'',style_fotos)
        worksheet.merge_range(24+conta3,9,63,9,'',style_fotos)
        worksheet.merge_range(24+conta3,10,63,10,'',style_fotos)
        worksheet.merge_range(24+conta3,11,63,11,'',style_fotos)
        worksheet.merge_range(24+conta3,12,63,12,'',style_fotos)

    #INSERTAR IMÁGENES: FOTOGRAFÍAS
    enlace1='C:\\Users\\HP\\Desktop\\DARYL_SUELOS\\Programas\\Puntos_Calicatas\\'+calicata+' ESP'+'.jpeg'
    enlace2='C:\\Users\\HP\\Desktop\\DARYL_SUELOS\\Programas\\Puntos_Calicatas\\'+calicata+' PROF'+'.jpeg'
    worksheet.insert_image('A12',enlace1,{'x_offset':60,'y_offset':10,'x_scale':0.14,'y_scale':0.12})
    worksheet.insert_image('A12',enlace2,{'x_offset':350,'y_offset':10,'x_scale':0.14,'y_scale':0.12})

    #INSERTAR IMÁGENES: SIMBOLOGÍA
    if sucs1!="-":
        enlace3='C:\\Users\\HP\\Desktop\\DARYL_SUELOS\\Programas\\Simbologia\\'+sucs1+'.jpg'
        escala1=conta1/21
        worksheet.insert_image(24,3,enlace3,{'x_offset':1.5, 'y_offset':1.5, 'y_scale':0.98*escala1, 'x_scale':0.62})
    if sucs2!="-":
        enlace4='C:\\Users\\HP\\Desktop\\DARYL_SUELOS\\Programas\\Simbologia\\'+sucs2+'.jpg'
        escala2=(conta2-conta1)/21
        worksheet.insert_image(24+conta1,3,enlace4,{'x_offset':1.5, 'y_offset':1.5, 'y_scale':0.98*escala2,'x_scale':0.62})
    if sucs3!="-":
        enlace5='C:\\Users\\HP\\Desktop\\DARYL_SUELOS\\Programas\\Simbologia\\'+sucs3+'.jpg'
        escala3=(conta3-conta2)/21
        worksheet.insert_image(24+conta2,3,enlace5,{'x_offset':1.5, 'y_offset':1.5, 'y_scale':1.01*escala3,'x_scale':0.62})

#Escala: 21 filas equivalen a escala 1.

workbook.close()

#Versión Final V1.0
#Créditos: Bach. Daryl Candia Nina
#Fecha: 12/07/2022 16:20
#Comentarios:
#- Las imágenes no abarcan toda la celda, pero su ubicación y tamaño es aceptable.